from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File
from fastapi.responses import JSONResponse
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime, timedelta
from collections import defaultdict
import json

from ..database import get_db
from ..models import (FlightSector, Aircraft, Airport, BlockTimeRule, TATRule,
                      Registration, Season, MaintenanceBlock, CalendarNote,
                      RouteColor, AppSetting)
from ..schemas import TimetableExportParams, ReportParams, ImportPayload

router = APIRouter()
import io

import openpyxl






# ── helpers ────────────────────────────────────────────────────────────────────
def time_to_minutes(hhmm: str) -> int:
    h, m = map(int, hhmm.split(":"))
    return h * 60 + m


def minutes_to_hhmm(m: int) -> str:
    m = m % 1440
    return f"{m // 60:02d}:{m % 60:02d}"


def block_minutes(dep: str, arr: str) -> int:
    d, a = time_to_minutes(dep), time_to_minutes(arr)
    if a < d:
        a += 1440
    return a - d


def apply_tz(hhmm: str, offset: float) -> str:
    mins = time_to_minutes(hhmm) + int(offset * 60)
    return minutes_to_hhmm(mins)


def lct_flight_date(sector_flight_date: str, dep_utc: str, tz_offset: float) -> str:
    """Compute the local-time (LCT) date of departure.
    If dep_utc + tz_offset rolls past midnight, the LCT date is the next day."""
    dep_min = time_to_minutes(dep_utc) + int(tz_offset * 60)
    dt = datetime.strptime(sector_flight_date, "%Y-%m-%d")
    if dep_min >= 1440:
        dt += timedelta(days=1)
    elif dep_min < 0:
        dt -= timedelta(days=1)
    return dt.strftime("%Y-%m-%d")


def fmt_display(sector: FlightSector, timezone: str, airports: dict) -> dict:
    dep_ap = airports.get(sector.origin, None)
    arr_ap = airports.get(sector.destination, None)
    orig_tz = dep_ap.timezone_offset if dep_ap else 7.0
    dest_tz = arr_ap.timezone_offset if arr_ap else 7.0

    if timezone == "UTC":
        dep_disp = sector.dep_utc
        arr_disp = sector.arr_utc
    else:
        dep_disp = apply_tz(sector.dep_utc, orig_tz)
        arr_disp = apply_tz(sector.arr_utc, dest_tz)

    # Day of week: Mon=1 .. Sun=7
    dt = datetime.strptime(sector.flight_date, "%Y-%m-%d")
    dow = dt.isoweekday()  # Monday=1, Sunday=7

    return {
        "id": sector.id,
        "aircraft_id": sector.aircraft_id,
        "flight_date": sector.flight_date,
        "origin": sector.origin,
        "destination": sector.destination,
        "route": f"{sector.origin}-{sector.destination}",
        "dep_utc": sector.dep_utc,
        "arr_utc": sector.arr_utc,
        "dep_display": dep_disp,
        "arr_display": arr_disp,
        "timezone": timezone,
        "dep_tz_label": f"+{int(orig_tz)}" if orig_tz >= 0 else str(int(orig_tz)),
        "arr_tz_label": f"+{int(dest_tz)}" if dest_tz >= 0 else str(int(dest_tz)),
        "block_time_minutes": block_minutes(sector.dep_utc, sector.arr_utc),
        "flight_number": sector.flight_number,
        "status": sector.status,
        "day_of_week": str(dow),
    }


def compress_dates(date_strings: List[str]) -> str:
    """Convert a sorted list of YYYY-MM-DD strings to a compact date-range string."""
    if not date_strings:
        return ""
    dates = sorted(date_strings)
    ranges = []
    start = dates[0]
    prev = dates[0]
    for d in dates[1:]:
        prev_dt = datetime.strptime(prev, "%Y-%m-%d")
        cur_dt = datetime.strptime(d, "%Y-%m-%d")
        if (cur_dt - prev_dt).days == 1:
            prev = d
        else:
            ranges.append((start, prev))
            start = d
            prev = d
    ranges.append((start, prev))

    parts = []
    for s, e in ranges:
        s_fmt = datetime.strptime(s, "%Y-%m-%d").strftime("%d%b%y").upper()
        e_fmt = datetime.strptime(e, "%Y-%m-%d").strftime("%d%b%y").upper()
        if s == e:
            parts.append(s_fmt)
        else:
            parts.append(f"{s_fmt}-{e_fmt}")
    return ", ".join(parts)


# ── Timetable export ───────────────────────────────────────────────────────────
@router.post("/timetable")
def export_timetable(params: TimetableExportParams, db: Session = Depends(get_db)):
    # When LCT mode, expand query window by 1 day before (a UTC sector from the
    # previous day may have an LCT departure that falls into the requested range).
    q_start = params.period_start
    q_end = params.period_end
    if params.timezone == "LCT":
        q_start = (datetime.strptime(params.period_start, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")

    # Exclude TẠM aircraft (registration_id IS NULL) from timetable
    tam_ids = {ac.id for ac in db.query(Aircraft).filter(Aircraft.registration_id.is_(None)).all()}

    sectors = (
        db.query(FlightSector)
        .filter(
            FlightSector.flight_date >= q_start,
            FlightSector.flight_date <= q_end,
            FlightSector.status == "active",
        )
        .order_by(FlightSector.flight_date, FlightSector.aircraft_id, FlightSector.dep_utc)
        .all()
    )
    sectors = [s for s in sectors if s.aircraft_id not in tam_ids]

    airports = {ap.code: ap for ap in db.query(Airport).all()}
    aircraft_map = {ac.id: ac for ac in db.query(Aircraft).all()}

    # Build registration map for seats lookup
    reg_map = {reg.id: reg for reg in db.query(Registration).all()}

    # When LCT, post-filter so only sectors whose LCT departure date is in range
    if params.timezone == "LCT":
        filtered = []
        for s in sectors:
            dep_ap = airports.get(s.origin, None)
            tz_off = dep_ap.timezone_offset if dep_ap else 7.0
            lct_date = lct_flight_date(s.flight_date, s.dep_utc, tz_off)
            if params.period_start <= lct_date <= params.period_end:
                filtered.append(s)
        sectors = filtered

    rows = [fmt_display(s, params.timezone, airports) for s in sectors]
    # Override flight_date and day_of_week with LCT date when in LCT mode
    if params.timezone == "LCT":
        for i, s in enumerate(sectors):
            dep_ap = airports.get(s.origin, None)
            tz_off = dep_ap.timezone_offset if dep_ap else 7.0
            ld = lct_flight_date(s.flight_date, s.dep_utc, tz_off)
            rows[i]["flight_date"] = ld
            dt = datetime.strptime(ld, "%Y-%m-%d")
            rows[i]["day_of_week"] = str(dt.isoweekday())

    for r in rows:
        ac = aircraft_map.get(r["aircraft_id"])
        r["aircraft_reg"] = ac.registration if ac else "?"
        r["line_order"] = ac.line_order if ac else 0
        # Seats from linked registration
        reg = reg_map.get(ac.registration_id) if ac and ac.registration_id else None
        r["seats"] = reg.seats if reg else 0

    if params.mode == "daily":
        # Merge identical sectors: group by (aircraft_id, route, dep_utc, arr_utc, flight_number)
        merged: dict = defaultdict(lambda: {"dates": [], "dows": set()})
        for r in rows:
            key = (
                r["aircraft_id"],
                r["route"],
                r["dep_utc"],
                r["arr_utc"],
                r.get("flight_number") or "",
            )
            merged[key]["dates"].append(r["flight_date"])
            merged[key]["dows"].add(int(r["day_of_week"]))
            merged[key]["row"] = r  # keep latest row as template

        merged_rows = []
        for key, grp in merged.items():
            r = dict(grp["row"])  # clone template row
            dows_sorted = sorted(grp["dows"])
            r["day_of_week"] = "".join(str(d) for d in dows_sorted)
            r["date_range"] = compress_dates(grp["dates"])
            r["flight_count"] = len(grp["dates"])
            r["total_seats"] = r["seats"] * r["flight_count"]
            merged_rows.append(r)

        # Default sort: aircraft (line_order) → flight_date → displayed dep time → route
        merged_rows.sort(key=lambda r: (
            (aircraft_map[r["aircraft_id"]].line_order if r["aircraft_id"] in aircraft_map else 0),
            r["aircraft_reg"],
            r.get("date_range", r.get("flight_date", "")),
            r["dep_display"],
            r["route"],
        ))
        return {"mode": "daily", "timezone": params.timezone, "rows": merged_rows}

    # GROUP mode – group by (origin, destination, dep_display [UTC hour:min key])
    grouped: dict = defaultdict(lambda: {"dates": [], "aircraft": set(), "dows": set(), "total_seats": 0, "min_line_order": 9999})
    for r in rows:
        key = f"{r['origin']}-{r['destination']}|{r['dep_utc']}|{r['arr_utc']}"
        grouped[key]["dates"].append(r["flight_date"])
        grouped[key]["aircraft"].add(r["aircraft_reg"])
        grouped[key]["dows"].add(int(r["day_of_week"]))
        grouped[key]["total_seats"] += r["seats"]
        grouped[key]["min_line_order"] = min(grouped[key]["min_line_order"], r.get("line_order", 0))
        grouped[key].update({
            "origin": r["origin"],
            "destination": r["destination"],
            "route": r["route"],
            "dep_utc": r["dep_utc"],
            "arr_utc": r["arr_utc"],
            "dep_display": r["dep_display"],
            "arr_display": r["arr_display"],
            "block_time_minutes": r["block_time_minutes"],
            "timezone": params.timezone,
        })

    group_rows = []
    for key, grp in grouped.items():
        # Compute DAY string: sorted unique DOWs, e.g. "1234567" for daily
        dows_sorted = sorted(grp["dows"])
        day_str = "".join(str(d) for d in dows_sorted)
        group_rows.append({
            "origin": grp["origin"],
            "destination": grp["destination"],
            "route": grp["route"],
            "dep_display": grp["dep_display"],
            "arr_display": grp["arr_display"],
            "dep_utc": grp["dep_utc"],
            "arr_utc": grp["arr_utc"],
            "block_time_minutes": grp["block_time_minutes"],
            "date_range": compress_dates(grp["dates"]),
            "flight_count": len(grp["dates"]),
            "aircraft": sorted(grp["aircraft"]),
            "day_of_week": day_str,
            "timezone": params.timezone,
            "total_seats": grp["total_seats"],
            "line_order": grp["min_line_order"],
        })

    group_rows.sort(key=lambda r: (r["origin"], r["destination"], r["dep_utc"]))

    return {"mode": "group", "timezone": params.timezone, "rows": group_rows}


# ── Report ─────────────────────────────────────────────────────────────────────
@router.post("/report")
def export_report(params: ReportParams, db: Session = Depends(get_db)):
    # When LCT mode, expand query window by 1 day before
    q_start = params.period_start
    q_end = params.period_end
    if params.timezone == "LCT":
        q_start = (datetime.strptime(params.period_start, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")

    # Exclude TẠM aircraft (registration_id IS NULL) from report
    tam_ids = {ac.id for ac in db.query(Aircraft).filter(Aircraft.registration_id.is_(None)).all()}

    sectors = (
        db.query(FlightSector)
        .filter(
            FlightSector.flight_date >= q_start,
            FlightSector.flight_date <= q_end,
            FlightSector.status == "active",
        )
        .all()
    )
    sectors = [s for s in sectors if s.aircraft_id not in tam_ids]

    airports = {ap.code: ap for ap in db.query(Airport).all()}
    aircraft_list = [ac for ac in db.query(Aircraft).order_by(Aircraft.line_order, Aircraft.id).all() if ac.id not in tam_ids]
    aircraft_map = {ac.id: ac for ac in aircraft_list}
    reg_map = {reg.id: reg for reg in db.query(Registration).all()}

    # When LCT, post-filter by LCT departure date
    if params.timezone == "LCT":
        filtered = []
        for s in sectors:
            dep_ap = airports.get(s.origin, None)
            tz_off = dep_ap.timezone_offset if dep_ap else 7.0
            lct_date = lct_flight_date(s.flight_date, s.dep_utc, tz_off)
            if params.period_start <= lct_date <= params.period_end:
                filtered.append(s)
        sectors = filtered

    # Compute block hours per aircraft
    per_aircraft: dict = defaultdict(lambda: {"block_minutes": 0, "sector_count": 0})
    per_route: dict = defaultdict(lambda: {"block_minutes": 0, "sector_count": 0, "dates": set(), "total_seats": 0})

    # Domestic / International breakdown
    dom_bm = 0
    dom_sectors = 0
    dom_seats = 0
    intl_bm = 0
    intl_sectors = 0
    intl_seats = 0
    dom_flight_dates: set = set()
    intl_flight_dates: set = set()
    per_aircraft_dom_bm: dict = defaultdict(int)
    per_aircraft_intl_bm: dict = defaultdict(int)

    for s in sectors:
        bt = block_minutes(s.dep_utc, s.arr_utc)
        per_aircraft[s.aircraft_id]["block_minutes"] += bt
        per_aircraft[s.aircraft_id]["sector_count"] += 1
        route_key = f"{s.origin}-{s.destination}"
        per_route[route_key]["block_minutes"] += bt
        per_route[route_key]["sector_count"] += 1
        # Seats for this sector
        ac = aircraft_map.get(s.aircraft_id)
        reg = reg_map.get(ac.registration_id) if ac and ac.registration_id else None
        sector_seats = reg.seats if reg else 0
        per_route[route_key]["total_seats"] += sector_seats
        # Use LCT date for unique-dates count when in LCT mode
        if params.timezone == "LCT":
            dep_ap = airports.get(s.origin, None)
            tz_off = dep_ap.timezone_offset if dep_ap else 7.0
            per_route[route_key]["dates"].add(lct_flight_date(s.flight_date, s.dep_utc, tz_off))
        else:
            per_route[route_key]["dates"].add(s.flight_date)

        # Classify domestic vs international
        orig_ap = airports.get(s.origin)
        dest_ap = airports.get(s.destination)
        is_dom = (orig_ap and orig_ap.is_domestic) and (dest_ap and dest_ap.is_domestic)
        # Determine flight date for day counting
        if params.timezone == "LCT":
            dep_ap_tz = airports.get(s.origin, None)
            tz_off_day = dep_ap_tz.timezone_offset if dep_ap_tz else 7.0
            f_date = lct_flight_date(s.flight_date, s.dep_utc, tz_off_day)
        else:
            f_date = s.flight_date
        if is_dom:
            dom_bm += bt
            dom_sectors += 1
            dom_seats += sector_seats
            dom_flight_dates.add(f_date)
            per_aircraft_dom_bm[s.aircraft_id] += bt
        else:
            intl_bm += bt
            intl_sectors += 1
            intl_seats += sector_seats
            intl_flight_dates.add(f_date)
            per_aircraft_intl_bm[s.aircraft_id] += bt

    # Build aircraft rows
    period_days = max(
        1,
        (datetime.strptime(params.period_end, "%Y-%m-%d") - datetime.strptime(params.period_start, "%Y-%m-%d")).days + 1,
    )

    aircraft_rows = []
    for ac in aircraft_list:
        data = per_aircraft.get(ac.id, {"block_minutes": 0, "sector_count": 0})
        bm = data["block_minutes"]
        reg = reg_map.get(ac.registration_id) if ac.registration_id else None
        ac_seats = reg.seats if reg else 0
        aircraft_rows.append({
            "line_order": ac.line_order,
            "aircraft_id": ac.id,
            "registration": ac.registration,
            "name": ac.name or "",
            "total_block_hours": round(bm / 60, 2),
            "total_block_minutes": bm,
            "sector_count": data["sector_count"],
            "avg_daily_block_hours": round(bm / 60 / period_days, 2),
            "seats": ac_seats,
            "total_seats": ac_seats * data["sector_count"],
        })

    # Averages
    total_bm = sum(r["total_block_minutes"] for r in aircraft_rows)
    avg_bh = round(total_bm / 60 / max(1, len(aircraft_rows)), 2)
    grand_total_seats = sum(r["total_seats"] for r in aircraft_rows)
    grand_total_sectors = sum(r["sector_count"] for r in aircraft_rows)

    # Route rows
    route_rows = []
    for rk, data in per_route.items():
        orig, dest = rk.split("-", 1)
        route_rows.append({
            "route": rk,
            "origin": orig,
            "destination": dest,
            "total_block_hours": round(data["block_minutes"] / 60, 2),
            "sector_count": data["sector_count"],
            "unique_dates": len(data["dates"]),
            "total_seats": data["total_seats"],
        })

    def _pair_sort_key(r):
        pair = "-".join(sorted([r["origin"], r["destination"]]))
        outbound = 0 if r["origin"] < r["destination"] else 1
        return (pair, outbound)
    route_rows.sort(key=_pair_sort_key)

    # Domestic / International breakdown for summary
    num_ac = max(1, len(aircraft_rows))
    dom_intl_breakdown = {
        "domestic": {
            "total_block_hours": round(dom_bm / 60, 2),
            "total_sectors": dom_sectors,
            "total_seats": dom_seats,
            "avg_per_aircraft_block_hours": round(dom_bm / 60 / num_ac, 2),
            "flight_days": len(dom_flight_dates),
        },
        "international": {
            "total_block_hours": round(intl_bm / 60, 2),
            "total_sectors": intl_sectors,
            "total_seats": intl_seats,
            "avg_per_aircraft_block_hours": round(intl_bm / 60 / num_ac, 2),
            "flight_days": len(intl_flight_dates),
        },
    }

    if params.sort_by == "route":
        return {
            "period_start": params.period_start,
            "period_end": params.period_end,
            "timezone": params.timezone,
            "period_days": period_days,
            "sort_by": "route",
            "route_rows": route_rows,
            "summary": {
                "total_block_hours": round(total_bm / 60, 2),
                "avg_per_aircraft_block_hours": avg_bh,
                "total_seats": grand_total_seats,
                "total_sectors": grand_total_sectors,
                "breakdown": dom_intl_breakdown,
            },
        }

    return {
        "period_start": params.period_start,
        "period_end": params.period_end,
        "timezone": params.timezone,
        "period_days": period_days,
        "sort_by": "aircraft",
        "aircraft_rows": aircraft_rows,
        "summary": {
            "total_block_hours": round(total_bm / 60, 2),
            "avg_per_aircraft_block_hours": avg_bh,
            "total_seats": grand_total_seats,
            "total_sectors": grand_total_sectors,
            "breakdown": dom_intl_breakdown,
        },
    }


# ── Full schedule export (JSON for file download/import) ────────────────────
@router.get("/schedule")
def export_schedule(db: Session = Depends(get_db)):
    airports = [{"code": a.code, "name": a.name, "timezone_offset": a.timezone_offset,
                 "curfew_open": a.curfew_open, "curfew_close": a.curfew_close}
                for a in db.query(Airport).all()]
    aircraft = [{"id": a.id, "registration": a.registration, "name": a.name,
                 "ac_type": a.ac_type, "line_order": a.line_order,
                 "color": a.color, "registration_id": a.registration_id}
                for a in db.query(Aircraft).order_by(Aircraft.line_order).all()]
    sectors = [
        {
            "id": s.id, "aircraft_id": s.aircraft_id, "flight_date": s.flight_date,
            "origin": s.origin, "destination": s.destination,
            "dep_utc": s.dep_utc, "arr_utc": s.arr_utc,
            "flight_number": s.flight_number, "status": s.status,
            "sequence": s.sequence, "color": s.color,
        }
        for s in db.query(FlightSector).all()
    ]
    bt_rules = [{"origin": r.origin, "destination": r.destination,
                 "block_time_minutes": r.block_time_minutes, "ats": r.ats}
                for r in db.query(BlockTimeRule).all()]
    tat_rules = [{"station": r.station, "min_tat_minutes": r.min_tat_minutes,
                  "is_domestic": r.is_domestic}
                 for r in db.query(TATRule).all()]
    registrations = [{"id": r.id, "registration": r.registration,
                      "aircraft_model": r.aircraft_model, "seats": r.seats,
                      "dw_type": r.dw_type, "mtow": r.mtow}
                     for r in db.query(Registration).all()]
    seasons = [{"name": s.name, "season_type": s.season_type, "year": s.year,
                "start_date": s.start_date, "end_date": s.end_date}
               for s in db.query(Season).all()]
    maintenance = [{"aircraft_id": m.aircraft_id, "label": m.label,
                    "start_date": m.start_date, "end_date": m.end_date,
                    "start_time": m.start_time, "end_time": m.end_time,
                    "color": m.color}
                   for m in db.query(MaintenanceBlock).all()]
    calendar_notes = [{"note_date": n.note_date, "note_end_date": n.note_end_date,
                       "start_time": n.start_time, "end_time": n.end_time,
                       "content": n.content, "color": n.color}
                      for n in db.query(CalendarNote).all()]
    route_colors = [{"origin": rc.origin, "destination": rc.destination, "color": rc.color}
                    for rc in db.query(RouteColor).all()]
    app_settings = [{"key": s.key, "value": s.value}
                    for s in db.query(AppSetting).all()]

    return {
        "version": "2.0",
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "airports": airports,
        "registrations": registrations,
        "aircraft": aircraft,
        "sectors": sectors,
        "block_time_rules": bt_rules,
        "tat_rules": tat_rules,
        "seasons": seasons,
        "maintenance": maintenance,
        "calendar_notes": calendar_notes,
        "route_colors": route_colors,
        "app_settings": app_settings,
    }


# ── Import ─────────────────────────────────────────────────────────────────────
@router.post("/import")
def import_schedule(payload: dict = Body(...), db: Session = Depends(get_db)):
    replace_all = payload.get("replace_all", False)

    if replace_all:
        # Delete in FK-safe order: sectors/maintenance first, then aircraft, then registrations
        db.query(FlightSector).delete()
        db.query(MaintenanceBlock).delete()
        db.query(Aircraft).delete()
        db.query(Registration).delete()
        db.query(BlockTimeRule).delete()
        db.query(TATRule).delete()
        db.query(Season).delete()
        db.query(CalendarNote).delete()
        db.query(RouteColor).delete()
        db.query(AppSetting).delete()
        db.commit()

    # Airports (always merge, never delete)
    for ap_data in payload.get("airports", []):
        code = ap_data["code"].upper()
        existing = db.query(Airport).filter(Airport.code == code).first()
        if not existing:
            db.add(Airport(code=code, name=ap_data.get("name", ""),
                           timezone_offset=ap_data.get("timezone_offset", 7.0),
                           curfew_open=ap_data.get("curfew_open"),
                           curfew_close=ap_data.get("curfew_close")))
        else:
            existing.name = ap_data.get("name") or existing.name
            existing.timezone_offset = ap_data.get("timezone_offset") if ap_data.get("timezone_offset") is not None else existing.timezone_offset
            if ap_data.get("curfew_open") is not None:
                existing.curfew_open = ap_data["curfew_open"]
            if ap_data.get("curfew_close") is not None:
                existing.curfew_close = ap_data["curfew_close"]
    db.commit()

    # Registrations (must come before aircraft for registration_id mapping)
    old_to_new_reg: dict = {}  # old_id → new_id
    for reg_data in payload.get("registrations", []):
        old_id = reg_data.get("id")
        existing = db.query(Registration).filter(
            Registration.registration == reg_data["registration"]).first()
        if not existing:
            new_reg = Registration(
                registration=reg_data["registration"],
                aircraft_model=reg_data.get("aircraft_model", ""),
                seats=reg_data.get("seats", 0),
                dw_type=reg_data.get("dw_type"),
                mtow=reg_data.get("mtow"),
            )
            db.add(new_reg)
            db.flush()
            if old_id is not None:
                old_to_new_reg[old_id] = new_reg.id
        else:
            existing.aircraft_model = reg_data.get("aircraft_model") or existing.aircraft_model
            existing.seats = reg_data.get("seats") if reg_data.get("seats") is not None else existing.seats
            if reg_data.get("dw_type") is not None:
                existing.dw_type = reg_data["dw_type"]
            if reg_data.get("mtow") is not None:
                existing.mtow = reg_data["mtow"]
            if old_id is not None:
                old_to_new_reg[old_id] = existing.id
    db.commit()

    # Aircraft (build id mapping from old→new)
    old_to_new_ac: dict = {}  # old_id → new_id
    for ac_data in payload.get("aircraft", []):
        old_id = ac_data.get("id")
        # Map registration_id through old→new mapping
        old_reg_id = ac_data.get("registration_id")
        new_reg_id = old_to_new_reg.get(old_reg_id, old_reg_id) if old_reg_id else None

        existing = db.query(Aircraft).filter(
            Aircraft.registration == ac_data["registration"]).first()
        if not existing:
            new_ac = Aircraft(
                registration=ac_data["registration"],
                name=ac_data.get("name"),
                ac_type=ac_data.get("ac_type"),
                line_order=ac_data.get("line_order", 0),
                color=ac_data.get("color"),
                registration_id=new_reg_id,
            )
            db.add(new_ac)
            db.flush()
            if old_id is not None:
                old_to_new_ac[old_id] = new_ac.id
        else:
            # Only overwrite fields that are explicitly provided and non-None
            if ac_data.get("name") is not None:
                existing.name = ac_data["name"]
            if ac_data.get("ac_type") is not None:
                existing.ac_type = ac_data["ac_type"]
            if ac_data.get("line_order") is not None:
                existing.line_order = ac_data["line_order"]
            if ac_data.get("color") is not None:
                existing.color = ac_data["color"]
            if new_reg_id is not None:
                existing.registration_id = new_reg_id
            if old_id is not None:
                old_to_new_ac[old_id] = existing.id
    db.commit()

    # Sectors (map aircraft_id through old→new)
    for s_data in payload.get("sectors", []):
        old_ac_id = s_data.get("aircraft_id")
        ac_id = old_to_new_ac.get(old_ac_id, old_ac_id)
        origin = s_data["origin"].upper()
        dest   = s_data["destination"].upper()
        # In merge mode, skip duplicates
        if not replace_all:
            dup = db.query(FlightSector).filter(
                FlightSector.aircraft_id == ac_id,
                FlightSector.flight_date == s_data["flight_date"],
                FlightSector.origin == origin,
                FlightSector.destination == dest,
                FlightSector.dep_utc == s_data["dep_utc"],
            ).first()
            if dup:
                continue
        db.add(FlightSector(
            aircraft_id=ac_id,
            flight_date=s_data["flight_date"],
            origin=origin,
            destination=dest,
            dep_utc=s_data["dep_utc"],
            arr_utc=s_data["arr_utc"],
            flight_number=s_data.get("flight_number"),
            status=s_data.get("status", "active"),
            sequence=s_data.get("sequence", 0),
            color=s_data.get("color"),
        ))
    db.commit()

    # Block-time rules
    for r_data in payload.get("block_time_rules", []):
        orig, dest = r_data["origin"].upper(), r_data["destination"].upper()
        existing = db.query(BlockTimeRule).filter(
            BlockTimeRule.origin == orig, BlockTimeRule.destination == dest).first()
        if existing:
            existing.block_time_minutes = r_data["block_time_minutes"]
            if r_data.get("ats") is not None:
                existing.ats = r_data["ats"]
        else:
            db.add(BlockTimeRule(origin=orig, destination=dest,
                                 block_time_minutes=r_data["block_time_minutes"],
                                 ats=r_data.get("ats")))

    # TAT rules
    for r_data in payload.get("tat_rules", []):
        station = r_data["station"].upper()
        existing = db.query(TATRule).filter(TATRule.station == station).first()
        if existing:
            if r_data.get("min_tat_minutes") is not None:
                existing.min_tat_minutes = r_data["min_tat_minutes"]
            if r_data.get("is_domestic") is not None:
                existing.is_domestic = r_data["is_domestic"]
        else:
            db.add(TATRule(station=station,
                           min_tat_minutes=r_data.get("min_tat_minutes", 40),
                           is_domestic=r_data.get("is_domestic")))

    # Seasons
    for s_data in payload.get("seasons", []):
        existing = db.query(Season).filter(
            Season.name == s_data["name"], Season.year == s_data["year"]).first()
        if not existing:
            db.add(Season(name=s_data["name"], season_type=s_data["season_type"],
                          year=s_data["year"], start_date=s_data["start_date"],
                          end_date=s_data["end_date"]))

    # Maintenance blocks (map aircraft_id)
    for m_data in payload.get("maintenance", []):
        old_ac_id = m_data.get("aircraft_id")
        ac_id = old_to_new_ac.get(old_ac_id, old_ac_id)
        db.add(MaintenanceBlock(
            aircraft_id=ac_id, label=m_data.get("label", "Maintenance"),
            start_date=m_data["start_date"], end_date=m_data["end_date"],
            start_time=m_data.get("start_time"), end_time=m_data.get("end_time"),
            color=m_data.get("color", "#f59e0b"),
        ))

    # Calendar notes
    for n_data in payload.get("calendar_notes", []):
        db.add(CalendarNote(
            note_date=n_data["note_date"],
            note_end_date=n_data.get("note_end_date"),
            start_time=n_data.get("start_time"),
            end_time=n_data.get("end_time"),
            content=n_data["content"],
            color=n_data.get("color", "#3b82f6"),
        ))

    # Route colors
    for rc_data in payload.get("route_colors", []):
        orig, dest = rc_data["origin"].upper(), rc_data["destination"].upper()
        existing = db.query(RouteColor).filter(
            RouteColor.origin == orig, RouteColor.destination == dest).first()
        if not existing:
            db.add(RouteColor(origin=orig, destination=dest, color=rc_data["color"]))
        else:
            existing.color = rc_data["color"]

    # App settings
    for s_data in payload.get("app_settings", []):
        existing = db.query(AppSetting).filter(AppSetting.key == s_data["key"]).first()
        if existing:
            existing.value = s_data.get("value")
        else:
            db.add(AppSetting(key=s_data["key"], value=s_data.get("value")))

    db.commit()

    cnt_sectors = len(payload.get("sectors", []))
    cnt_aircraft = len(payload.get("aircraft", []))
    return {"ok": True, "message": f"Import successful: {cnt_aircraft} aircraft, {cnt_sectors} sectors"}
#-- help excel----
def _write_sheet(wb: openpyxl.Workbook, title: str, headers: list[str], rows: list[dict]):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def _read_sheet(wb: openpyxl.Workbook, title: str) -> list[dict]:
    if title not in wb.sheetnames:
        return []
    ws = wb[title]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # skip empty rows
        result.append(dict(zip(headers, row)))
    return result


# ── Export ─────────────────────────────────────────────────────────────────────

@router.get("/schedulexlsx")
def export_schedule(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Airports
    airports = db.query(Airport).all()
    _write_sheet(wb, "airports",
                 ["code", "name", "timezone_offset", "curfew_open", "curfew_close"],
                 [{"code": a.code, "name": a.name, "timezone_offset": a.timezone_offset,
                   "curfew_open": a.curfew_open, "curfew_close": a.curfew_close}
                  for a in airports])

    # Registrations
    registrations = db.query(Registration).all()
    _write_sheet(wb, "registrations",
                 ["id", "registration", "aircraft_model", "seats", "dw_type", "mtow"],
                 [{"id": r.id, "registration": r.registration,
                   "aircraft_model": r.aircraft_model, "seats": r.seats,
                   "dw_type": r.dw_type, "mtow": r.mtow}
                  for r in registrations])

    # Aircraft
    aircraft = db.query(Aircraft).order_by(Aircraft.line_order).all()
    _write_sheet(wb, "aircraft",
                 ["id", "registration", "name", "ac_type", "line_order", "color", "registration_id"],
                 [{"id": a.id, "registration": a.registration, "name": a.name,
                   "ac_type": a.ac_type, "line_order": a.line_order,
                   "color": a.color, "registration_id": a.registration_id}
                  for a in aircraft])

    # Sectors
    sectors = db.query(FlightSector).all()
    _write_sheet(wb, "sectors",
                 ["id", "aircraft_id", "flight_date", "origin", "destination",
                  "dep_utc", "arr_utc", "flight_number", "status", "sequence", "color"],
                 [{"id": s.id, "aircraft_id": s.aircraft_id, "flight_date": s.flight_date,
                   "origin": s.origin, "destination": s.destination,
                   "dep_utc": s.dep_utc, "arr_utc": s.arr_utc,
                   "flight_number": s.flight_number, "status": s.status,
                   "sequence": s.sequence, "color": s.color}
                  for s in sectors])

    # Block-time rules
    bt_rules = db.query(BlockTimeRule).all()
    _write_sheet(wb, "block_time_rules",
                 ["origin", "destination", "block_time_minutes", "ats"],
                 [{"origin": r.origin, "destination": r.destination,
                   "block_time_minutes": r.block_time_minutes, "ats": r.ats}
                  for r in bt_rules])

    # TAT rules
    tat_rules = db.query(TATRule).all()
    _write_sheet(wb, "tat_rules",
                 ["station", "min_tat_minutes", "is_domestic"],
                 [{"station": r.station, "min_tat_minutes": r.min_tat_minutes,
                   "is_domestic": r.is_domestic}
                  for r in tat_rules])

    # Seasons
    seasons = db.query(Season).all()
    _write_sheet(wb, "seasons",
                 ["name", "season_type", "year", "start_date", "end_date"],
                 [{"name": s.name, "season_type": s.season_type, "year": s.year,
                   "start_date": s.start_date, "end_date": s.end_date}
                  for s in seasons])

    # Maintenance
    maintenance = db.query(MaintenanceBlock).all()
    _write_sheet(wb, "maintenance",
                 ["aircraft_id", "label", "start_date", "end_date", "start_time", "end_time", "color"],
                 [{"aircraft_id": m.aircraft_id, "label": m.label,
                   "start_date": m.start_date, "end_date": m.end_date,
                   "start_time": m.start_time, "end_time": m.end_time,
                   "color": m.color}
                  for m in maintenance])

    # Calendar notes
    calendar_notes = db.query(CalendarNote).all()
    _write_sheet(wb, "calendar_notes",
                 ["note_date", "note_end_date", "start_time", "end_time", "content", "color"],
                 [{"note_date": n.note_date, "note_end_date": n.note_end_date,
                   "start_time": n.start_time, "end_time": n.end_time,
                   "content": n.content, "color": n.color}
                  for n in calendar_notes])

    # Route colors
    route_colors = db.query(RouteColor).all()
    _write_sheet(wb, "route_colors",
                 ["origin", "destination", "color"],
                 [{"origin": rc.origin, "destination": rc.destination, "color": rc.color}
                  for rc in route_colors])

    # App settings
    app_settings = db.query(AppSetting).all()
    _write_sheet(wb, "app_settings",
                 ["key", "value"],
                 [{"key": s.key, "value": s.value} for s in app_settings])

    # Stream file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"schedule_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Import ─────────────────────────────────────────────────────────────────────

@router.post("/importxlsx")
async def import_schedule(
    file: UploadFile = File(...),
    replace_all: bool = False,
    db: Session = Depends(get_db),
):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    if replace_all:
        db.query(FlightSector).delete()
        db.query(MaintenanceBlock).delete()
        db.query(Aircraft).delete()
        db.query(Registration).delete()
        db.query(BlockTimeRule).delete()
        db.query(TATRule).delete()
        db.query(Season).delete()
        db.query(CalendarNote).delete()
        db.query(RouteColor).delete()
        db.query(AppSetting).delete()
        db.commit()

    # ── Airports ──────────────────────────────────────────────────────────────
    for ap in _read_sheet(wb, "airports"):
        code = str(ap["code"]).upper() if ap.get("code") else None
        if not code:
            continue
        existing = db.query(Airport).filter(Airport.code == code).first()
        if not existing:
            db.add(Airport(
                code=code,
                name=ap.get("name", ""),
                timezone_offset=ap.get("timezone_offset", 7.0),
                curfew_open=ap.get("curfew_open"),
                curfew_close=ap.get("curfew_close"),
            ))
        else:
            if ap.get("name"):
                existing.name = ap["name"]
            if ap.get("timezone_offset") is not None:
                existing.timezone_offset = ap["timezone_offset"]
            if ap.get("curfew_open") is not None:
                existing.curfew_open = ap["curfew_open"]
            if ap.get("curfew_close") is not None:
                existing.curfew_close = ap["curfew_close"]
    db.commit()

    # ── Registrations ─────────────────────────────────────────────────────────
    old_to_new_reg: dict = {}
    for reg in _read_sheet(wb, "registrations"):
        old_id = reg.get("id")
        reg_str = str(reg["registration"]) if reg.get("registration") else None
        if not reg_str:
            continue
        existing = db.query(Registration).filter(Registration.registration == reg_str).first()
        if not existing:
            new_reg = Registration(
                registration=reg_str,
                aircraft_model=reg.get("aircraft_model", ""),
                seats=int(reg["seats"]) if reg.get("seats") is not None else 0,
                dw_type=reg.get("dw_type"),
                mtow=reg.get("mtow"),
            )
            db.add(new_reg)
            db.flush()
            if old_id is not None:
                old_to_new_reg[int(old_id)] = new_reg.id
        else:
            if reg.get("aircraft_model"):
                existing.aircraft_model = reg["aircraft_model"]
            if reg.get("seats") is not None:
                existing.seats = int(reg["seats"])
            if reg.get("dw_type") is not None:
                existing.dw_type = reg["dw_type"]
            if reg.get("mtow") is not None:
                existing.mtow = reg["mtow"]
            if old_id is not None:
                old_to_new_reg[int(old_id)] = existing.id
    db.commit()

    # ── Aircraft ──────────────────────────────────────────────────────────────
    old_to_new_ac: dict = {}
    for ac in _read_sheet(wb, "aircraft"):
        old_id = ac.get("id")
        reg_str = str(ac["registration"]) if ac.get("registration") else None
        if not reg_str:
            continue
        old_reg_id = ac.get("registration_id")
        new_reg_id = old_to_new_reg.get(int(old_reg_id), int(old_reg_id)) if old_reg_id else None

        existing = db.query(Aircraft).filter(Aircraft.registration == reg_str).first()
        if not existing:
            new_ac = Aircraft(
                registration=reg_str,
                name=ac.get("name"),
                ac_type=ac.get("ac_type"),
                line_order=int(ac["line_order"]) if ac.get("line_order") is not None else 0,
                color=ac.get("color"),
                registration_id=new_reg_id,
            )
            db.add(new_ac)
            db.flush()
            if old_id is not None:
                old_to_new_ac[int(old_id)] = new_ac.id
        else:
            if ac.get("name") is not None:
                existing.name = ac["name"]
            if ac.get("ac_type") is not None:
                existing.ac_type = ac["ac_type"]
            if ac.get("line_order") is not None:
                existing.line_order = int(ac["line_order"])
            if ac.get("color") is not None:
                existing.color = ac["color"]
            if new_reg_id is not None:
                existing.registration_id = new_reg_id
            if old_id is not None:
                old_to_new_ac[int(old_id)] = existing.id
    db.commit()

    # ── Sectors ───────────────────────────────────────────────────────────────
    for s in _read_sheet(wb, "sectors"):
        old_ac_id = s.get("aircraft_id")
        ac_id = old_to_new_ac.get(int(old_ac_id), int(old_ac_id)) if old_ac_id else None
        origin = str(s["origin"]).upper()
        dest   = str(s["destination"]).upper()
        # In merge mode, skip sectors that already exist to avoid duplicates
        if not replace_all:
            dup = db.query(FlightSector).filter(
                FlightSector.aircraft_id == ac_id,
                FlightSector.flight_date == s["flight_date"],
                FlightSector.origin == origin,
                FlightSector.destination == dest,
                FlightSector.dep_utc == s["dep_utc"],
            ).first()
            if dup:
                continue
        db.add(FlightSector(
            aircraft_id=ac_id,
            flight_date=s["flight_date"],
            origin=origin,
            destination=dest,
            dep_utc=s["dep_utc"],
            arr_utc=s["arr_utc"],
            flight_number=s.get("flight_number"),
            status=s.get("status", "active"),
            sequence=int(s["sequence"]) if s.get("sequence") is not None else 0,
            color=s.get("color"),
        ))
    db.commit()

    # ── Block-time rules ──────────────────────────────────────────────────────
    for r in _read_sheet(wb, "block_time_rules"):
        orig = str(r["origin"]).upper()
        dest = str(r["destination"]).upper()
        existing = db.query(BlockTimeRule).filter(
            BlockTimeRule.origin == orig, BlockTimeRule.destination == dest).first()
        if existing:
            existing.block_time_minutes = r["block_time_minutes"]
            if r.get("ats") is not None:
                existing.ats = r["ats"]
        else:
            db.add(BlockTimeRule(origin=orig, destination=dest,
                                 block_time_minutes=r["block_time_minutes"],
                                 ats=r.get("ats")))

    # ── TAT rules ─────────────────────────────────────────────────────────────
    for r in _read_sheet(wb, "tat_rules"):
        station = str(r["station"]).upper()
        existing = db.query(TATRule).filter(TATRule.station == station).first()
        if existing:
            if r.get("min_tat_minutes") is not None:
                existing.min_tat_minutes = r["min_tat_minutes"]
            if r.get("is_domestic") is not None:
                existing.is_domestic = r["is_domestic"]
        else:
            db.add(TATRule(station=station,
                           min_tat_minutes=r.get("min_tat_minutes", 40),
                           is_domestic=r.get("is_domestic")))

    # ── Seasons ───────────────────────────────────────────────────────────────
    for s in _read_sheet(wb, "seasons"):
        existing = db.query(Season).filter(
            Season.name == s["name"], Season.year == s["year"]).first()
        if not existing:
            db.add(Season(name=s["name"], season_type=s.get("season_type"),
                          year=s["year"], start_date=s["start_date"],
                          end_date=s["end_date"]))

    # ── Maintenance ───────────────────────────────────────────────────────────
    for m in _read_sheet(wb, "maintenance"):
        old_ac_id = m.get("aircraft_id")
        ac_id = old_to_new_ac.get(int(old_ac_id), int(old_ac_id)) if old_ac_id else None
        db.add(MaintenanceBlock(
            aircraft_id=ac_id,
            label=m.get("label", "Maintenance"),
            start_date=m["start_date"],
            end_date=m["end_date"],
            start_time=m.get("start_time"),
            end_time=m.get("end_time"),
            color=m.get("color", "#f59e0b"),
        ))

    # ── Calendar notes ────────────────────────────────────────────────────────
    for n in _read_sheet(wb, "calendar_notes"):
        db.add(CalendarNote(
            note_date=n["note_date"],
            note_end_date=n.get("note_end_date"),
            start_time=n.get("start_time"),
            end_time=n.get("end_time"),
            content=n["content"],
            color=n.get("color", "#3b82f6"),
        ))

    # ── Route colors ──────────────────────────────────────────────────────────
    for rc in _read_sheet(wb, "route_colors"):
        orig = str(rc["origin"]).upper()
        dest = str(rc["destination"]).upper()
        existing = db.query(RouteColor).filter(
            RouteColor.origin == orig, RouteColor.destination == dest).first()
        if not existing:
            db.add(RouteColor(origin=orig, destination=dest, color=rc["color"]))
        else:
            existing.color = rc["color"]

    # ── App settings ──────────────────────────────────────────────────────────
    for s in _read_sheet(wb, "app_settings"):
        existing = db.query(AppSetting).filter(AppSetting.key == s["key"]).first()
        if existing:
            existing.value = s.get("value")
        else:
            db.add(AppSetting(key=s["key"], value=s.get("value")))

    db.commit()

    cnt_sectors = len(_read_sheet(wb, "sectors"))
    cnt_aircraft = len(_read_sheet(wb, "aircraft"))
    return {"ok": True, "message": f"Import successful: {cnt_aircraft} aircraft, {cnt_sectors} sectors"}