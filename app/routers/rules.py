from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import csv
import io
import datetime
import openpyxl

from ..database import get_db
from ..models import TATRule, BlockTimeRule, Airport, Registration
from ..schemas import (
    TATRuleCreate, TATRuleOut,
    BlockTimeRuleCreate, BlockTimeRuleOut,
    AirportCreate, AirportOut,
    RegistrationCreate, RegistrationUpdate, RegistrationOut,
)

router = APIRouter()


# ── Airports ───────────────────────────────────────────────────────────────────
@router.get("/airports", response_model=List[AirportOut])
def list_airports(db: Session = Depends(get_db)):
    return db.query(Airport).order_by(Airport.code).all()


@router.post("/airports", response_model=AirportOut, status_code=201)
def create_airport(payload: AirportCreate, db: Session = Depends(get_db)):
    existing = db.query(Airport).filter(Airport.code == payload.code.upper()).first()
    if existing:
        raise HTTPException(400, f"Airport '{payload.code}' already exists")
    ap = Airport(code=payload.code.upper(), name=payload.name, timezone_offset=payload.timezone_offset)
    db.add(ap)
    db.commit()
    db.refresh(ap)
    return ap


@router.put("/airports/{code}", response_model=AirportOut)
def update_airport(code: str, payload: AirportCreate, db: Session = Depends(get_db)):
    ap = db.query(Airport).filter(Airport.code == code.upper()).first()
    if not ap:
        raise HTTPException(404, "Airport not found")
    ap.name = payload.name
    ap.timezone_offset = payload.timezone_offset
    db.commit()
    db.refresh(ap)
    return ap


@router.delete("/airports/{code}", status_code=204)
def delete_airport(code: str, db: Session = Depends(get_db)):
    ap = db.query(Airport).filter(Airport.code == code.upper()).first()
    if not ap:
        raise HTTPException(404, "Airport not found")
    db.delete(ap)
    db.commit()


# ── TAT Rules ──────────────────────────────────────────────────────────────────
MASS_TAT_STATIONS = {"__DOMESTIC__", "__INTL__"}


@router.get("/tat", response_model=List[TATRuleOut])
def list_tat_rules(db: Session = Depends(get_db)):
    return db.query(TATRule).filter(TATRule.station.notin_(MASS_TAT_STATIONS)).order_by(TATRule.station).all()


@router.get("/tat/mass")
def get_mass_tat(db: Session = Depends(get_db)):
    """Return the mass (default) TAT rules for domestic and international."""
    dom = db.query(TATRule).filter(TATRule.station == "__DOMESTIC__").first()
    intl = db.query(TATRule).filter(TATRule.station == "__INTL__").first()
    return {
        "domestic": dom.min_tat_minutes if dom else 40,
        "international": intl.min_tat_minutes if intl else 60,
    }


@router.put("/tat/mass")
def set_mass_tat(payload: dict, db: Session = Depends(get_db)):
    """Save mass TAT defaults. Expects {domestic: int, international: int}."""
    for key, station in [("domestic", "__DOMESTIC__"), ("international", "__INTL__")]:
        minutes = payload.get(key)
        if minutes is None:
            continue
        existing = db.query(TATRule).filter(TATRule.station == station).first()
        if existing:
            existing.min_tat_minutes = int(minutes)
        else:
            db.add(TATRule(station=station, min_tat_minutes=int(minutes)))
    db.commit()
    return {"ok": True}


@router.post("/tat", response_model=TATRuleOut, status_code=201)
def create_tat_rule(payload: TATRuleCreate, db: Session = Depends(get_db)):
    existing = db.query(TATRule).filter(TATRule.station == payload.station.upper()).first()
    if existing:
        existing.min_tat_minutes = payload.min_tat_minutes
        existing.is_domestic = payload.is_domestic
        db.commit()
        db.refresh(existing)
        return existing
    rule = TATRule(
        station=payload.station.upper(),
        min_tat_minutes=payload.min_tat_minutes,
        is_domestic=payload.is_domestic,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return rule


@router.put("/tat/{rule_id}", response_model=TATRuleOut)
def update_tat_rule(rule_id: int, payload: TATRuleCreate, db: Session = Depends(get_db)):
    rule = db.query(TATRule).filter(TATRule.id == rule_id).first()
    if not rule:
        raise HTTPException(404, "TAT rule not found")
    rule.station = payload.station.upper()
    rule.min_tat_minutes = payload.min_tat_minutes
    rule.is_domestic = payload.is_domestic
    db.commit()
    db.refresh(rule)
    return rule


@router.delete("/tat/{rule_id}", status_code=204)
def delete_tat_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = db.query(TATRule).filter(TATRule.id == rule_id).first()
    if not rule:
        raise HTTPException(404, "TAT rule not found")
    db.delete(rule)
    db.commit()


# ── Block-time Rules ───────────────────────────────────────────────────────────
@router.get("/blocktime", response_model=List[BlockTimeRuleOut])
def list_block_time_rules(db: Session = Depends(get_db)):
    return db.query(BlockTimeRule).order_by(BlockTimeRule.origin, BlockTimeRule.destination).all()


@router.post("/blocktime", response_model=BlockTimeRuleOut, status_code=201)
def create_block_time_rule(payload: BlockTimeRuleCreate, db: Session = Depends(get_db)):
    orig = payload.origin.upper()
    dest = payload.destination.upper()
    existing = db.query(BlockTimeRule).filter(
        BlockTimeRule.origin == orig, BlockTimeRule.destination == dest
    ).first()
    if existing:
        raise HTTPException(409, f"Block time rule {orig}-{dest} đã tồn tại. Vui lòng sửa thay vì tạo mới.")
    rule = BlockTimeRule(origin=orig, destination=dest, block_time_minutes=payload.block_time_minutes, ats=payload.ats)
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return rule


@router.put("/blocktime/{rule_id}", response_model=BlockTimeRuleOut)
def update_block_time_rule(rule_id: int, payload: BlockTimeRuleCreate, db: Session = Depends(get_db)):
    rule = db.query(BlockTimeRule).filter(BlockTimeRule.id == rule_id).first()
    if not rule:
        raise HTTPException(404, "Block-time rule not found")
    orig = payload.origin.upper()
    dest = payload.destination.upper()
    # Check for duplicate: another rule with same origin-dest (different id)
    dup = db.query(BlockTimeRule).filter(
        BlockTimeRule.origin == orig,
        BlockTimeRule.destination == dest,
        BlockTimeRule.id != rule_id,
    ).first()
    if dup:
        raise HTTPException(409, f"Block time rule {orig}-{dest} đã tồn tại (ID {dup.id})")
    rule.origin = orig
    rule.destination = dest
    rule.block_time_minutes = payload.block_time_minutes
    rule.ats = payload.ats
    db.commit()
    db.refresh(rule)
    return rule


@router.delete("/blocktime/{rule_id}", status_code=204)
def delete_block_time_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = db.query(BlockTimeRule).filter(BlockTimeRule.id == rule_id).first()
    if not rule:
        raise HTTPException(404, "Block-time rule not found")
    db.delete(rule)
    db.commit()


# ── Excel Export/Import ────────────────────────────────────────────────────────
def minutes_to_hhmm(minutes: int) -> str:
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"


def minutes_to_decimal(minutes: int) -> str:
    h = minutes // 60
    frac = round((minutes % 60) / 60 * 100)
    return f"{h:02d}.{frac:02d}"


def hhmm_to_minutes(value) -> int:
    """Convert various representations of a time duration to total minutes.

    openpyxl may return a cell value as:
      - str  "01:30"  → 90 min
      - datetime.time  → hours*60 + minutes
      - float (Excel serial fraction of a day, e.g. 0.0625 = 1h30m) → *1440
    """
    if isinstance(value, datetime.time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)):
        # Excel stores time as fraction of 24h
        total_minutes = round(float(value) * 24 * 60)
        return total_minutes
    # fallback: treat as "HH:MM" string
    text = str(value).strip()
    parts = text.replace(",", ":").split(":")
    return int(parts[0]) * 60 + int(parts[1])


@router.get("/tat/export")
def export_tat_excel(db: Session = Depends(get_db)):
    rules = db.query(TATRule).order_by(TATRule.station).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TAT Rules"
    
    # Header
    ws.append(["Station", "Min TAT"])
    
    # Data
    for r in rules:
        ws.append([r.station, minutes_to_hhmm(r.min_tat_minutes)])
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tat_rules.xlsx"}
    )


@router.post("/tat/import")
async def import_tat_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        station = str(row[0]).upper().strip()
        time_str = str(row[1]).strip()
        minutes = hhmm_to_minutes(time_str)
        
        existing = db.query(TATRule).filter(TATRule.station == station).first()
        if existing:
            existing.min_tat_minutes = minutes
        else:
            db.add(TATRule(station=station, min_tat_minutes=minutes))
        imported += 1
    
    db.commit()
    return {"imported": imported}


@router.get("/blocktime/export")
def export_blocktime_excel(db: Session = Depends(get_db)):
    rules = db.query(BlockTimeRule).order_by(BlockTimeRule.origin, BlockTimeRule.destination).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Block Time Rules"
    
    # Header
    ws.append(["Origin", "Destination", "Block Time", "Decimal", "ATS"])
    
    # Data
    for r in rules:
        ws.append([r.origin, r.destination, minutes_to_hhmm(r.block_time_minutes), minutes_to_decimal(r.block_time_minutes), r.ats or ""])
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=blocktime_rules.xlsx"}
    )


@router.post("/blocktime/import")
async def import_blocktime_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        origin = str(row[0]).upper().strip()
        dest = str(row[1]).upper().strip()
        time_str = str(row[2]).strip()
        minutes = hhmm_to_minutes(time_str)
        ats = str(row[4]).strip() if len(row) > 4 and row[4] else None
        
        existing = db.query(BlockTimeRule).filter(
            BlockTimeRule.origin == origin,
            BlockTimeRule.destination == dest
        ).first()
        if existing:
            existing.block_time_minutes = minutes
            existing.ats = ats
        else:
            db.add(BlockTimeRule(origin=origin, destination=dest, block_time_minutes=minutes, ats=ats))
        imported += 1
    
    db.commit()
    return {"imported": imported}


# ── Registration ───────────────────────────────────────────────────────────────
@router.get("/registration", response_model=List[RegistrationOut])
def list_registrations(db: Session = Depends(get_db)):
    return db.query(Registration).order_by(Registration.registration).all()


@router.post("/registration", response_model=RegistrationOut, status_code=201)
def create_registration(payload: RegistrationCreate, db: Session = Depends(get_db)):
    reg = payload.registration.upper()
    existing = db.query(Registration).filter(Registration.registration == reg).first()
    if existing:
        raise HTTPException(400, f"Registration '{reg}' already exists")
    r = Registration(registration=reg, aircraft_model=payload.aircraft_model, seats=payload.seats, dw_type=payload.dw_type, mtow=payload.mtow)
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


@router.put("/registration/{reg_id}", response_model=RegistrationOut)
def update_registration(reg_id: int, payload: RegistrationUpdate, db: Session = Depends(get_db)):
    r = db.query(Registration).filter(Registration.id == reg_id).first()
    if not r:
        raise HTTPException(404, "Registration not found")
    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(r, field, value)
    db.commit()
    db.refresh(r)
    return r


@router.delete("/registration/{reg_id}", status_code=204)
def delete_registration(reg_id: int, db: Session = Depends(get_db)):
    r = db.query(Registration).filter(Registration.id == reg_id).first()
    if not r:
        raise HTTPException(404, "Registration not found")
    db.delete(r)
    db.commit()


@router.get("/registration/export/excel")
def export_registration_excel(db: Session = Depends(get_db)):
    regs = db.query(Registration).order_by(Registration.registration).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"

    ws.append(["Số đăng bạ", "Mẫu máy bay", "Số ghế", "D/W", "MTOW"])
    for r in regs:
        ws.append([r.registration, r.aircraft_model, r.seats, r.dw_type or "", r.mtow or ""])

    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=registrations.xlsx"},
    )


@router.get("/registration/export/csv")
def export_registration_csv(db: Session = Depends(get_db)):
    regs = db.query(Registration).order_by(Registration.registration).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Số đăng bạ", "Mẫu máy bay", "Số ghế", "D/W", "MTOW"])
    for r in regs:
        writer.writerow([r.registration, r.aircraft_model, r.seats, r.dw_type or "", r.mtow or ""])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=registrations.csv"},
    )


@router.post("/registration/import/excel")
async def import_registration_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        reg_code = str(row[0]).upper().strip()
        model = str(row[1]).strip() if row[1] else ""
        seats = int(row[2]) if row[2] else 0
        dw = str(row[3]).strip() if len(row) > 3 and row[3] else None
        mtow_val = float(row[4]) if len(row) > 4 and row[4] else None

        existing = db.query(Registration).filter(Registration.registration == reg_code).first()
        if existing:
            existing.aircraft_model = model
            existing.seats = seats
            existing.dw_type = dw if dw else None
            existing.mtow = mtow_val
        else:
            db.add(Registration(
                registration=reg_code, aircraft_model=model, seats=seats,
                dw_type=dw if dw else None, mtow=mtow_val,
            ))
        imported += 1

    db.commit()
    return {"imported": imported}
